import os.path
import zipfile
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook
import os

# Предусловие: в папке resources находится несколько разных файлов: pdf, xlsx, csv
# Task_1: Средствами Python запаковать в zip архив файлы разных типов: pdf, xlsx, csv
# Task_2: Поместить архив в папку /resources
# Task_3: Реализовать чтение и проверку содержимого каждого файла из архива

def test_create_archive():
    zip_archive = zipfile.ZipFile('resources/archive.zip', 'w')
    zip_archive.write('resources/file1.xlsx')
    zip_archive.write('resources/file2.csv')
    zip_archive.write('resources/Siegfrieds_Journey.pdf')
    zip_archive.close()


#pdf - checks text on page №7, number of sheets (232), and file size
def test_pdf():
    with zipfile.ZipFile('resources/archive.zip') as myzip:
        with myzip.open("resources/Siegfrieds_Journey.pdf") as pdf_data:
            pdf_data = PdfReader(pdf_data)
            page = pdf_data.pages[6]
            text = page.extract_text()
            assert "SIEGFRIED SASSOON\n(from theportrait byGlyn Philpot, R.A.)\n" in text
            number_of_page = len(pdf_data.pages)
            assert number_of_page == 232
            pdf_size = os.path.getsize('resources/Siegfrieds_Journey.pdf')
            assert pdf_size == 24799258 #(23,6 МБ (24 799 258 байт))

#csv - check with text ("01.01.1912") and file size
def test_read_and_assert_csv():
    with open('resources/file2.csv') as csvfile:
        csvfile = csv.reader(csvfile)
        for line_no, line in enumerate(csvfile, 0):
            if line_no == 1:
                assert '01.01.1912' in line[0]
                xlsx_size = os.path.getsize('resources/file2.csv')
                assert xlsx_size == 136  #  (136 байт)

#xlsx - check with text ("date_new") and file size
def test_xlsx_abc():
    with zipfile.ZipFile('resources/archive.zip') as myzip:
        with myzip.open("resources/file1.xlsx") as xlsx_data:
            workbook = load_workbook(xlsx_data)
            sheet = workbook.active
            sheet = sheet.cell(row=1, column=2).value
            assert "date_new" in sheet
            xlsx_size = os.path.getsize('resources/file1.xlsx')
            assert xlsx_size == 9147 #47,3 МБ (49 617 610 байт)

#xlsx - check with date
def test_xlsx_datatime():
    with zipfile.ZipFile('resources/archive.zip') as myzip:
        with myzip.open("resources/file1.xlsx") as xlsx_data:
            workbook = load_workbook(xlsx_data)
            datetime = workbook.active
            datetime = datetime.cell(row=4, column=1).value
            assert "03.01.1912" in datetime











